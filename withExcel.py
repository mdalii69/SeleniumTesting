from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

# Load the Excel sheet
workbook = load_workbook(r"C:\\Users\\abc\\VS Code\\Testing\\test_data.xlsx")  # Ensure file format is .xlsx
sheet = workbook.active

# Initialize the WebDriver
driver = webdriver.Chrome()
driver.maximize_window()

# Iterate through rows in the sheet, starting from the second row
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username, password = row[0], row[1]
    
    # Open the target application
    driver.get("https://www.saucedemo.com/")
    time.sleep(2)
    
    # Enter username
    driver.find_element(By.ID, "user-name").send_keys(username)
    # Enter password
    driver.find_element(By.ID, "password").send_keys(password)
    # Click login button
    driver.find_element(By.ID, "login-button").click()
    time.sleep(2)
    
    # Verify login and log out (if applicable)
    try:
        # Check if login was successful by finding logout button or profile
        driver.find_element(By.ID, "react-burger-menu-btn").click()
        time.sleep(1)
        driver.find_element(By.ID, "logout_sidebar_link").click()
        print(f"Login successful for: {username}")
    except Exception as e:
        print(f"Login failed for: {username}")
        print(f"Error: {e}")
    
    time.sleep(2)

# Close the browser
driver.quit()
